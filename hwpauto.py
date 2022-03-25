from tkinter import *
from tkinter import filedialog
import tkinter.ttk
import win32com.client as win32
import openpyxl

def save():
    savepath = filedialog.askdirectory()
    number=1
    hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
    hwp.XHwpWindows.Item(0).Visible = True
    hwp.Open(hwppath)
    defaultName = fileName[0].get()
    igenName =  fileName[1].get()
    colums_data_list = []
    for i in range(len(hwpField)):
        colums_data_list.append(get_column_data(hwpField[i]))
        
    print(colums_data_list)
    

    for j in range(len(ws['A'])-1):  ### 행 반복
        for i in range(len(hwpField)):  ### 열 반복
            hwp.PutFieldText(hwpField[i], " ")   
        for i in range(len(hwpField)):
            hwp.PutFieldText(hwpField[i], colums_data_list[i][j])   
        if pdfCheck.get() == 1:
            hwp.SaveAs(savepath+"/"+str(number)+"."+str(defaultName)+"_"+str(get_column_data(igenName)[j])+".pdf" , "PDF")
        else:
            hwp.SaveAs(savepath+"/"+str(number)+"."+str(defaultName)+"_"+str(get_column_data(igenName)[j])+".hwp" , "HWP")
        number=number+1
        
    print('save')


def xlspathSelct():
    global excelpath, df, table,datarows, pdfCheck, fileName, ws 
    fileName =[]
    rows = []
    excelpath = filedialog.askopenfilename()
    xlsPathEntry.delete(0,END)
    xlsPathEntry.insert(0,excelpath)
    wb = openpyxl.load_workbook(filename=excelpath)
    ws = wb.active
    print(excelpath)
    #print(list(ws))
    rows=get_row_data('1')
    print(rows)
    
    xlsFieldEntry.insert(0,list(rows))
    
    """
    for i in columnlength:
        colunms.append(ws.cell(row=1, column = i))
    print(rowlegnth)
    """
   
    ####

    lastplace = [50,130]
    defaultNameLabel = Label(window, text = "기본 파일 이름")
    defaultNameLabel.place(x=50, y = lastplace[1]+50, width = 80, height = 30)
    defaultNameEntry = Entry(window, width =250)
    defaultNameEntry.place(x=140,y=lastplace[1]+50, width =350, height = 30)
    
    igenNameLabel = Label(window, text = "파일 고유 이름")
    igenNameLabel.place(x=50, y = lastplace[1]+100, width = 80, height = 30)
    igenNameEntry = Entry(window, width =100)
    igenNameEntry.place(x=140,y=lastplace[1]+100, width =80, height = 30)
    
    saveTypeLabel = Label(window, text = "pdf 저장")
    saveTypeLabel.place(x=50, y = lastplace[1]+150, width = 80, height = 30)
    pdfCheck = IntVar()
    pdfCheckbox = Checkbutton(window, text = "PDF", variable = pdfCheck)
    pdfCheckbox.place(x=150, y = lastplace[1]+150)

    saveButton = Button(window, text = "저장", command = save)
    saveButton.place(x=300-75,y=lastplace[1]+180,  width = 150, height = 30)
    
    fileName.append(defaultNameEntry) 
    fileName.append(igenNameEntry)
    
def get_column_letter(columnName):
    for i in range(len(ws['1'])):
        if ws.cell(row=1, column = i+1).value == columnName:
            return ws.cell(row=1, column = i+1).column_letter
                   
def get_row_data(rowNum):
    rows=[]
    for i in range(len(ws[rowNum])):
        rows.append(ws.cell(row=int(rowNum), column = i+1).value)
    #print(rows)
    return rows

    
def get_column_data(columName):
    columns=[]
    colums_letter = get_column_letter(columName)
    print(colums_letter)
    for i in range(len(ws[str(colums_letter)])):
        columns.append(ws[str(colums_letter)+str(i+1)].value)
    columns.remove(columName)
    return columns
    
    

def hwppathSelct():
    global hwppath, hwpField
    hwppath = filedialog.askopenfilename()
    hwpPathEntry.delete(0,END)
    hwpPathEntry.insert(0,hwppath)
    print(hwppath)
    hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
    hwp.XHwpWindows.Item(0).Visible = True
    hwp.Open(hwppath)
    hwpField = hwp.GetFieldList(0,None).split("\x02")
    print(hwpField)
    hwpFieldEntry.insert(0,list(hwpField))
    #fields = "한글파일 누름틀 : " + str(hwpField)
    #hwpFieldEntry.insert(0,fields)


window = tkinter.Tk()
window.title("한글파일 자동입력 프로그램(by Song)")
window.geometry("600x500")
#window.resizable(False,False)
table = 0

##### 위젯 설정 ######

hwpLabel = Label(window, text = "한글 양식 파일")
xlsLabel = Label(window, text = "엑셀 파일")

hwpPathEntry = Entry(window, width =300)
xlsPathEntry = Entry(window, width =300)

hwpPathselectButton =Button(window, text = "찾기", command = hwppathSelct)
xlsselectButton = Button(window, text = "찾기", command = xlspathSelct)

#hwpFieldEntry = Entry(window, width =350)
#xlsFieldEntry = Entry(window, width =350)


##### 위젯 배치 ########


hwpLabel.place(x=50,y=30, width = 80, height = 30)
xlsLabel.place(x=50,y=60, width = 80, height = 30)

hwpPathEntry.place(x=130,y=30, width =300, height = 30)
xlsPathEntry.place(x=130,y=60, width =300, height = 30)

hwpPathselectButton.place(x=460,y=30,  width =50,height = 30)
xlsselectButton.place(x=460,y=60,  width =50,height = 30)

hwpFieldLabel = Label(window, text = "한글 누름틀")
hwpFieldLabel.place(x=50,y=110, width = 80, height = 30)

xlsFieldLabel = Label(window, text = "엑셀 열 이름")
xlsFieldLabel.place(x=50,y=140, width = 80, height = 30)

hwpFieldEntry = Entry(window, width =350)
xlsFieldEntry = Entry(window, width =350)
hwpFieldEntry.place(x=130,y=110, width =350, height = 30)
xlsFieldEntry.place(x=130,y=140, width =350, height = 30)



window.mainloop()

